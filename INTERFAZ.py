import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import calendar

# ...existing code...

def create_excel_calendar(year, month, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]}"

    # Set the title font
    title_font = Font(size=14, bold=True)

    # Set the headers for the days of the week
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    cal = calendar.Calendar(firstweekday=0)

    # Fill in the days of the month
    col_num = 1
    for week in cal.monthdayscalendar(year, month):
        for day in week:
            col_letter = get_column_letter(col_num)
            if day != 0:
                day_name = days[(col_num - 1) % 7]
                ws[f"{col_letter}1"] = day_name
                ws[f"{col_letter}1"].font = title_font
                ws[f"{col_letter}2"] = day
                col_num += 1

    # Save the workbook
    wb.save(filename)

# Example usage
create_excel_calendar(2025, 1, r"\\celerrabr\Sesiones\D7627S023\Comum_S023\2.- OPERACIONES\ARIADNA\Calendar.xlsx")

# ...existing code...
